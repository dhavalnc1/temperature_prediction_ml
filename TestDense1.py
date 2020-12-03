# https://www.kaggle.com/ternaryrealm/lstm-time-series-explorations-with-keras

import math
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

# Original data set retrieved from here:
# https://datamarket.com/data/set/22u3/international-airline-passengers-monthly-totals-in-thousands-jan-49-dec-60#!ds=22u3&display=line

# data = pd.read_csv("D:\\UNCW\\Capstone\\ML1.csv",
#                       usecols = [0],
#                       engine = "python",
#                       skipfooter = 3)


data = pd.read_excel("D:\\UNCW\\Capstone\\InputData\\1.xlsx",
                      usecols=[1])

# Print some data rows.
print(data.head())

# Create a time series plot.
plt.figure(figsize = (15, 5))
plt.plot(data, label = "Temperature")
plt.xlabel("Time Index")
plt.ylabel("Temperature in F")
plt.title("Temperature data of Sensor")
plt.legend()
# print(plt.show())


# Let's load the required libs.
# We'll be using the Tensorflow backend (default).
from keras.models import Sequential
from keras.layers.recurrent import LSTM
from keras.layers.core import Dense, Activation, Dropout
from sklearn.preprocessing import MinMaxScaler
from sklearn.metrics import mean_squared_error
from sklearn.utils import shuffle

# Get the raw data values from the pandas data frame.
data_raw = data.values.astype("float32")

# We apply the MinMax scaler from sklearn
# to normalize data in the (0, 1) interval.
scaler = MinMaxScaler(feature_range = (0, 1))
dataset = scaler.fit_transform(data_raw)

# Print a few values.
print(dataset[0:5])

TRAIN_SIZE = 0.60

train_size = int(len(dataset) * TRAIN_SIZE)
test_size = len(dataset) - train_size
train, test = dataset[0:train_size, :], dataset[train_size:len(dataset), :]
print("Number of entries (training set, test set): " + str((len(train), len(test))))


# FIXME: This helper function should be rewritten using numpy's shift function. See below.
def create_dataset(dataset, window_size):
    data_X, data_Y = [], []
    for i in range(len(dataset) - window_size - 1):
        a = dataset[i:(i + window_size), 0]
        data_X.append(a)
        data_Y.append(dataset[i + window_size, 0])
    return(np.array(data_X), np.array(data_Y))


# Create test and training sets for one-step-ahead regression.
window_size = 50
train_X, train_Y = create_dataset(train, window_size)
test_X, test_Y = create_dataset(test, window_size)

print("train_X shape 0 ", train_X.shape)
print("train_Y shape 0 ", train_Y.shape)

print(train[0:5])
print("Y = ", train_Y[0:5])

print("Original training data shape:")
print(train_X.shape)
print("Train X")
print(train_X)

print("Train Y")
print(train_Y)

# Reshape the input data into appropriate form for Keras.
train_X = np.reshape(train_X, (train_X.shape[0], 1, train_X.shape[1]))
test_X = np.reshape(test_X, (test_X.shape[0], 1, test_X.shape[1]))
print("New training data shape:")
print(train_X.shape)


def fit_model(train_X, train_Y, window_size):
    model = Sequential()

    model.add(LSTM(4,
                   input_shape=(1, window_size)))
    model.add(Dense(1))
    model.compile(loss="mean_squared_error",
                  optimizer="adam")
    model.fit(train_X,
              train_Y,
              epochs=100,
              batch_size=500,
              verbose=2)

    return (model)


# Fit the first model.
model1 = fit_model(train_X, train_Y, window_size)


def predict_and_score(model, X, Y):
    # Make predictions on the original scale of the data.
    pred = scaler.inverse_transform(model.predict(X))
    # Prepare Y data to also be on the original scale for interpretability.
    orig_data = scaler.inverse_transform([Y])
    # Calculate RMSE.
    score = math.sqrt(mean_squared_error(orig_data[0], pred[:, 0]))
    return(score, pred)


print("train_X shape ", train_X.shape)
print("train_Y shape ", train_Y.shape)
rmse_train, train_predict = predict_and_score(model1, train_X, train_Y)
rmse_test, test_predict = predict_and_score(model1, test_X, test_Y)

print("Training data score: %.2f RMSE" % rmse_train)
print("Test data score: %.2f RMSE" % rmse_test)

# Start with training predictions.
train_predict_plot = np.empty_like(dataset)
train_predict_plot[:, :] = np.nan
train_predict_plot[window_size:len(train_predict) + window_size, :] = train_predict

# Add test predictions.
test_predict_plot = np.empty_like(dataset)
test_predict_plot[:, :] = np.nan
test_predict_plot[len(train_predict) + (window_size * 2) + 1:len(dataset) - 1, :] = test_predict

# Create the plot.
plt.figure(figsize = (15, 5))
plt.plot(scaler.inverse_transform(dataset), label = "True value")
plt.plot(train_predict_plot, label = "Training set prediction")
plt.plot(test_predict_plot, label = "Test set prediction")
plt.xlabel("Time Index")
plt.ylabel("Temp in F")
plt.title("Comparison true vs. predicted training / test")
plt.legend()
print(plt.show())

import openpyxl

#Create workbook object
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
sheet.title='Sheet #1'

#Add titles in the first row of each column
# sheet.cell(row=1, column=1).value='Train X'
# sheet.cell(row=1, column=2).value='Train Y'
# sheet.cell(row=1, column=3).value='Train Pred'
# sheet.cell(row=1, column=4).value='Test X'
# sheet.cell(row=1, column=5).value='Test Y'
sheet.cell(row=1, column=6).value='Test Pred'

#Loop to set the value of each cell
for i in range(0, len(test_predict)):
    sheet.cell(row=i+2, column=6).value=test_predict[i]
    # sheet.cell(row=i+2, column=1).value=train_X[i]
    # sheet.cell(row=i+2, column=2).value=train_Y[i]
    # sheet.cell(row=i+2, column=3).value=train_predict[i]
    # sheet.cell(row=i+2, column=4).value=test_X[i]
    # sheet.cell(row=i+2, column=5).value=test_Y[i]



#Finally, save the file and give it a name
wb.save('Dense 1.xlsx')