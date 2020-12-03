# https://www.kaggle.com/ternaryrealm/lstm-time-series-explorations-with-keras
# Test model with different input/output size, windows/output different
# Test model with different params, loss function, optimizer
# Combine model 1 and model 2
# model 1 = predict temperature
# model 2 = find failure in predicted temperature
# Plot model2 prediction vs real data
# Date: 11/30

import math
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
# Let's load the required libs.
# We'll be using the Tensorflow backend (default).
from keras.models import Sequential
from keras.layers.recurrent import LSTM
from keras.layers.core import Dense, Activation, Dropout
from sklearn.preprocessing import MinMaxScaler
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import mean_squared_error
from openpyxl import Workbook

# The location of source file
sourceFile = "D:\\UNCW\\Capstone\\InputData\\3.xlsx"

# Read temperature data of sensor from Excel file
Tempdata = pd.read_excel(sourceFile, usecols=[1])

# Read failure data of sensor from Excel file
Failuredata = pd.read_excel(sourceFile, usecols=[3])

# Parameters tuning
window_size = 336  # input window, number of input values
output_size = 48  # output window, number of predictions
neuron_count = 32  # number of nuerons to be used for training
batch_count = 50  # batch size
epoch_count = 100  # number of epoch

# Print some data rows.
print("Tempdata.head(10)")
print(Tempdata.head(10))

print("Failuredata.head(10)")
print(Failuredata.head(10))

# Get the raw data values from the pandas data frame.
temperature_data_raw = Tempdata.values.astype("float32")
failure_data_raw = Failuredata.values.astype("float32")

# We apply the MinMax scaler from sklearn
# to normalize data in the (0, 1) interval.
# scaler = MinMaxScaler(feature_range=(0, 1))
scaler = StandardScaler()
temperature_dataset = scaler.fit_transform(temperature_data_raw)
failure_dataset = scaler.fit_transform(failure_data_raw)

# split temperature data into train/test size
TRAIN_SIZE = 0.80
train_size = int(len(temperature_dataset) * TRAIN_SIZE)
test_size = len(temperature_dataset) - train_size
train, test = temperature_dataset[0:train_size, :], temperature_dataset[train_size:len(temperature_dataset), :]
print("Number of entries (training set, test set): " + str((len(train), len(test))))


def create_temperature_dataset(datain, input_size, output_size):
    data_X, data_Y = [], []
    for i in range(len(datain) - input_size - output_size - 1):
        a = datain[i:(i + input_size), 0]
        data_X.append(a)
        b = datain[(i+input_size):(i + input_size + output_size), 0]
        data_Y.append(b)
    return np.array(data_X), np.array(data_Y)


# Create temperature test and training sets for one-step-ahead regression.
train_temp_X, train_temp_Y = create_temperature_dataset(train, window_size, output_size)
test_temp_X, test_temp_Y = create_temperature_dataset(test, window_size, output_size)


# Reshape the input data into appropriate form for Keras.
train_temp_X = np.reshape(train_temp_X, (train_temp_X.shape[0], 1, train_temp_X.shape[1]))
test_temp_X = np.reshape(test_temp_X, (test_temp_X.shape[0], 1, test_temp_X.shape[1]))
print("New training data shape X:")
print(train_temp_X.shape)
print("New testing data shape X:")
print(test_temp_X.shape)


def fit_model(train_X, train_Y, window_size, output_size):
    model = Sequential()
    model.add(LSTM(neuron_count, input_shape=(1, window_size)))
    model.add(Dense(output_size))
    model.compile(loss="mean_squared_error",
                  optimizer="adam")
    model.fit(train_X,
              train_Y,
              epochs=epoch_count,
              batch_size=batch_count,
              verbose=2)

    return model


# Fit the first model.
# model1 = temperature model
# model2 = failure model
model1 = fit_model(train_temp_X, train_temp_Y, window_size, output_size)


# serialize model to JSON
model_json = model1.to_json()
with open("model.json", "w") as json_file:
    json_file.write(model_json)
# serialize weights to HDF5
model1.save_weights("temperatureModel.h5")
print("Saved temperature model to disk")


# function to predict and score input data, i am not using it for now
def predict_and_score(model, X, Y):
    # Make predictions on the original scale of the data.
    pred = scaler.inverse_transform(model.predict(X))
    # Prepare Y data to also be on the original scale for interpretability.
    orig_data = scaler.inverse_transform([Y])
    # Calculate RMSE.
    score = math.sqrt(mean_squared_error(orig_data[0], pred[:, 0]))
    return score, pred


# print("Test Input shape:", test_temp_X.shape)
predTemp = scaler.inverse_transform(model1.predict(test_temp_X))
# print("Predicted Test Output Shape:", predX.shape)
OriginalTestY = scaler.inverse_transform(test_temp_Y)


# empty rmseValue array
# rmseValue = []

# print("OriginalTestY 0 = ", OriginalTestY[0])
# print("predX 0 = ", predX[0])

# loop to print real data vs predicted data and RMSE
# for i in range(200):
#     rmseValue.append(round(np.sqrt(((OriginalTestY[i] - np.round(predX[i], 1)) ** 2).mean()), 2))


# print("RMSE: Min ", min(rmseValue), " Max ", max(rmseValue), " Avg ", round((sum(rmseValue)/(len(rmseValue))), 2))
# Create workbook object
# wb = Workbook()
# sheet = wb.active
# sheet.title = 'Sheet #1'

# Add titles in the first row of each column
# sheet.cell(row=1, column=1).value = 'Output 1'
# sheet.cell(row=1, column=2).value = 'Output 2'
# sheet.cell(row=1, column=3).value = 'Output 3'
# sheet.cell(row=1, column=4).value = 'Output 4'
# sheet.cell(row=1, column=5).value = 'Output 5'
# sheet.cell(row=1, column=(output_size + 1)).value = 'RMSE'

# Loop to set the value of each cell
# for i in range(200):  # (0, len(test_predict)):
#     for j in range(output_size):
#         sheet.cell(row=i+2, column=j+1).value = round(predX[i][j], 2)
#     sheet.cell(row=i + 2, column=(output_size + 1)).value = rmseValue[i]


# # #Finally, save the file and give it a name
# wb.save('Dense 3.xlsx')


# ############################################################
# Train/Test failure model
# ############################################################

temperature_dataset = scaler.fit_transform(temperature_data_raw)
failure_dataset = scaler.fit_transform(failure_data_raw)


# split failure data into train/test size
# TRAIN_SIZE = 0.80
# train_size = int(len(failure_dataset) * TRAIN_SIZE)
# test_size = len(failure_dataset) - train_size
# train, test = failure_dataset[0:train_size, :], failure_dataset[train_size:len(failure_dataset), :]
# print("Number of entries Failure (training set, test set): " + str((len(train), len(test))))


# Prepare failure data for input
def create_failure_dataset(datain, dataout, input_size):
    data_X, data_Y = [], []
    for i in range(len(datain) - input_size - 1):
        a = datain[i:(i + input_size), 0]
        data_X.append(a)
        data_Y.append(dataout[i, 0])
        # data_Y.append(dataset[i + window_size, 0])
    return np.array(data_X), np.array(data_Y)


# Create failure test and training sets for one-step-ahead regression.
input_size_fail = 1
output_size_fail = 1
train_fail_X, train_fail_Y = create_failure_dataset(temperature_dataset, failure_dataset, input_size_fail)
train_fail_Y = failure_dataset[0:len(failure_dataset)-2]
# test_fail_X, test_fail_Y = create_failure_dataset(test, window_size, output_size)


# Reshape the input data into appropriate form for Keras.
train_fail_X = np.reshape(train_fail_X, (train_fail_X.shape[0], 1, train_fail_X.shape[1]))
# test_fail_X = np.reshape(test_fail_X, (test_fail_X.shape[0], 1, test_fail_X.shape[1]))
print("New training failure data shape X:")
print(train_fail_X.shape)
# print("New testing data shape X:")
# print(test_fail_X.shape)

# Fit the second model.
# model1 = temperature model
# model2 = failure model
model2 = fit_model(train_fail_X, train_fail_Y, input_size_fail, output_size_fail)

# serialize model to JSON
model_json = model2.to_json()
with open("model.json", "w") as json_file:
    json_file.write(model_json)
# serialize weights to HDF5
model1.save_weights("failModel.h5")
print("Saved failure model to disk")


# Calculate RMSE of trained model(model 2) for failure prediction
rmse_train, train_predict = predict_and_score(model2, train_fail_X, train_fail_Y)
# print("Training data score: %.2f RMSE" % rmse_train)
print("Train failure data score: %.2f RMSE" % rmse_train)

# Scale-back predicted value
# trainPredictScaled = scaler.fit_transform(train_predict)
trainPredictScaled = train_predict

# Start with training predictions.
train_predict_plot = np.empty_like(failure_dataset)
train_predict_plot[:, :] = np.nan
train_predict_plot[input_size_fail:len(trainPredictScaled) + input_size_fail, :] = trainPredictScaled
# train_predict_plot[window_size:len(train_predict) + window_size, :] = train_predict
# Create the plot.
plt.figure(figsize=(15, 5))
plt.plot(scaler.inverse_transform(failure_dataset), label = "True value")
# plt.plot(failure_dataset, label = "True value")
plt.plot(train_predict_plot, label = "Training set prediction")
# plt.plot(test_predict_plot, label = "Test set prediction")
plt.xlabel("Time Index")
plt.ylabel("Failure")
plt.title("Comparison true vs. predicted training / test")
plt.legend()
print(plt.show())